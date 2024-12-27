# 资产的service层
import json
import os
import shutil
import uuid
from io import BytesIO

import pandas as pd
from datetime import datetime

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side
from typing_extensions import assert_type

from api.model.assets import AssetCreateApiModel, AssetFlowApiModel
from db.models.asset.models import AssetBasicInfo, AssetManufacturesInfo, AssetPartsInfo, AssetPositionsInfo, \
    AssetContractsInfo, AssetCustomersInfo, AssetBelongsInfo, AssetType, AssetFlowsInfo, AssetManufactureRelationInfo
from db.models.asset.sql import AssetSQL
from math import ceil
from oslo_log import log

from utils.constant import ASSET_SERVER_TEMPLATE_FILE_DIR, asset_equipment_columns, asset_basic_info_columns, \
    asset_manufacture_info_columns, asset_position_info_columns, asset_contract_info_columns, asset_belong_info_columns, \
    asset_customer_info_columns, asset_part_info_columns, asset_network_basic_info_columns, \
    asset_network_manufacture_info_columns, asset_network_position_info_columns, asset_network_basic_info_extra_columns, \
    ASSET_NETWORK_TEMPLATE_FILE_DIR, ASSET_NETWORK_FLOW_TEMPLATE_FILE_DIR, asset_network_flow_info_columns
from utils.datetime import change_excel_date_to_timestamp

LOG = log.getLogger(__name__)

# 定义边框样式
thin_border = Border(
    left=Side(border_style="thin", color="000000"),  # 左边框
    right=Side(border_style="thin", color="000000"),  # 右边框
    top=Side(border_style="thin", color="000000"),  # 上边框
    bottom=Side(border_style="thin", color="000000")  # 下边框
)

class AssetsService:

    # 查询资产列表
    def list_assets(self, asset_id, asset_ids, asset_name, asset_category, asset_type, asset_status, frame_position, cabinet_position, u_position, equipment_number, asset_number, sn_number, department_name, user_name, page, page_size, sort_keys, sort_dirs):
        # 业务逻辑
        try:
            # 按照条件从数据库中查询数据
            count, data = AssetSQL.list_asset(asset_id, asset_ids, asset_name, asset_category, asset_type, asset_status, frame_position, cabinet_position, u_position, equipment_number, asset_number, sn_number, department_name, user_name, page, page_size, sort_keys, sort_dirs)
            # 数据处理
            ret = []
            # 遍历
            for r in data:
                # 填充数据
                temp = {}
                temp["asset_id"] = r.id
                temp["asset_type_id"] = r.asset_type_id
                temp["asset_category"] = r.asset_category
                temp["asset_type"] = r.asset_type
                temp["asset_type_name_zh"] = r.asset_type_name_zh
                temp["asset_name"] = r.name
                temp["equipment_number"] = r.equipment_number
                temp["sn_number"] = r.sn_number
                temp["asset_number"] = r.asset_number
                temp["asset_status"] = r.asset_status
                temp["asset_description"] = r.description
                temp["extra"] = r.extra
                # 厂商信息
                temp_manufacture = {}
                temp_manufacture["id"] = r.manufacture_id
                temp_manufacture["name"] = r.manufacture_name
                temp_manufacture["description"] = r.manufacture_description
                temp_manufacture["extra"] = r.manufacture_extra
                temp["asset_manufacturer"] = temp_manufacture
                # 位置信息
                temp_position = {}
                temp_position["id"] = r.position_id
                temp_position["frame_position"] = r.position_frame_position
                temp_position["cabinet_position"] = r.position_cabinet_position
                temp_position["u_position"] = r.position_u_position
                temp_position["description"] = r.position_description
                temp["asset_position"] = temp_position
                # 合同信息
                temp_contract = {}
                temp_contract["id"] = r.contract_id
                temp_contract["contract_number"] = r.contract_number
                temp_contract["purchase_date"] = None if r.contract_purchase_date is None else r.contract_purchase_date.timestamp() * 1000
                temp_contract["batch_number"] = r.contract_batch_number
                temp_contract["description"] = r.contract_description
                temp["asset_contract"] = temp_contract
                # 归属信息
                temp_belong = {}
                temp_belong["id"] = r.belong_id
                temp_belong["department_id"] = r.belong_department_id
                temp_belong["department_name"] = r.belong_department_name
                temp_belong["user_id"] = r.belong_user_id
                temp_belong["user_name"] = r.belong_user_name
                temp_belong["tel_number"] = r.belong_tel_number
                temp_belong["description"] = r.belong_contract_description
                temp["asset_belong"] = temp_belong
                # 租户信息
                temp_cutomer = {}
                temp_cutomer["id"] = r.customer_id
                temp_cutomer["customer_id"] = r.customer_customer_id
                temp_cutomer["customer_name"] = r.customer_customer_name
                temp_cutomer["rental_duration"] = r.customer_rental_duration
                temp_cutomer["start_date"] = None if r.customer_start_date is None else r.customer_start_date.timestamp() * 1000
                temp_cutomer["end_date"] = None if r.customer_end_date is None else r.customer_end_date.timestamp() * 1000
                temp_cutomer["vlan_id"] = r.customer_vlan_id
                temp_cutomer["float_ip"] = r.customer_float_ip
                temp_cutomer["band_width"] = r.customer_band_width
                temp_cutomer["description"] = r.customer_description
                temp["asset_customer"] = temp_cutomer
                # 配件信息
                temp["asset_part"] = self.list_assets_parts(r.id)
                # 流量信息 列表上不需要
                # temp["asset_flow"] = self.list_assets_flows(r.id)
                # 加入列表
                ret.append(temp)

            # 返回数据
            res = {}
            # 页数相关信息
            if page and page_size:
                res['currentPage'] = page
                res['pageSize'] = page_size
                res['totalPages'] = ceil(count / int(page_size))
            res['total'] = count
            res['data'] = ret
            return res
        except Exception as e:
            import traceback
            traceback.print_exc()
            return None



# 返回对应数据
#     return {
#         "flag": True,
#         "errorCode": None,
#         "errorMessage": None,
#         "resData": {
#             "currentPage": 1,
#             "pageSize": 10,
#             "total": count,
#             "totalPages": 2,
#             "data":
#                 [
#                     {
#                         "uuid":"5a1c563b91fd4cbfaf518cd6e8a42059",
#                         "asset_name":"服务器",
#                         "asset_description":"服务器描述",
#                     }
#                 ]
#         }
#     }



    def create_asset(self, asset):
        # 业务校验 todo
        if asset is None:
            return None
        # 资产id
        asset_id = None
        try:
            # 数据校验
            # 1、名称空
            if asset.asset_name is None or asset.asset_type_id is None:
                raise Exception
            # 2、重名
            count, _ = AssetSQL.list_asset(None,None, asset.asset_name, None,None, None, None, None, None, None, None, None, None, None, 1,10,None,None)
            if count > 0:
                LOG.error("asset name exist")
                raise Exception
            # 3. 查询资产类型
            asset_type_list = AssetSQL.list_asset_type(asset.asset_type_id, None, None, None)
            if not asset_type_list:
                LOG.error("asset type not exist")
                raise Exception
            asset_type_name = asset_type_list[0].asset_type_name
            # 数据组装
            # 1、资产基础信息
            asset_basic_info_db = self.convert_asset_basic_info_db(asset)
            asset_basic_info_db.asset_type = asset_type_name
            asset_basic_info_db.asset_category = asset_type_name.split("_")[0]
            # 资产的id重新生成覆盖
            asset.asset_id = asset_basic_info_db.id
            asset_id = asset_basic_info_db.id
            # 2、资产厂商信息 资产关联的厂商可能已经存在也可能不存在
            asset_manufacture_info_db = None
            asset_manufacture_info_db_new = None
            if asset.asset_manufacturer:
                asset_manufacture_info_db = self.check_manufacturer_exists(asset)
            # 数据库不存在当前厂商
            if not asset_manufacture_info_db:
                asset_manufacture_info_db_new = self.convert_asset_manufacturer_info_db(asset)
            # 建立资产与厂商关联关系对象
            asset_manufacture_relation_info_db = self.convert_manufacturer_relation_info_db(asset, asset_manufacture_info_db, asset_manufacture_info_db_new)
            # 3、资产位置信息
            asset_position_info_db = self.convert_asset_position_info_db(asset)
            # 4、资产合同信息
            asset_contract_info_db = self.convert_asset_contract_info_db(asset)
            # 5、资产归属信息
            asset_belong_info_db = self.convert_asset_belong_info_db(asset)
            # 6、资产租赁信息
            asset_customer_info_db = self.convert_asset_customer_info_db(asset)
            # 7、资产配件信息
            asset_part_info_db = self.convert_asset_part_info_db(asset)
            # 8、资产流量信息
            asset_flow_info_db = None # self.convert_asset_flow_info_db_from_asset(asset)
            # 保存对象
            AssetSQL.create_asset(asset_basic_info_db, asset_manufacture_info_db_new, asset_manufacture_relation_info_db, asset_position_info_db, asset_contract_info_db, asset_belong_info_db, asset_customer_info_db, asset_part_info_db, asset_flow_info_db)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e
        # 成功返回资产id
        return asset_id


    def check_manufacturer_exists(self, asset):
        # 默认空
        asset_manufacture_info_db = None
        # 首先根据条件查询
        if asset.asset_manufacturer:
            # 先根据id查询
            if asset.asset_manufacturer.id:
                return AssetSQL.get_manufacture_by_id(asset.asset_manufacturer.id)
            if asset.asset_manufacturer.name:
                return AssetSQL.get_manufacture_by_name(asset.asset_manufacturer.name)
        # 返回
        return asset_manufacture_info_db


    # 资产创建时基础对象数据转换
    def convert_asset_basic_info_db(self, asset):
        # 数据转化为db对象
        asset_basic_info_db = AssetBasicInfo(
            id=uuid.uuid4().hex,
            asset_type_id=asset.asset_type_id,
            asset_category=asset.asset_category,
            asset_type=asset.asset_type,
            name=asset.asset_name,
            description=asset.asset_description,
            equipment_number=asset.equipment_number,
            sn_number=asset.sn_number,
            asset_number=asset.asset_number,
            asset_status=asset.asset_status,
            extra=json.dumps(asset.extra) if asset.extra else asset.extra
        )
        # 返回数据
        return asset_basic_info_db


    # 资产创建时厂商对象数据转换
    def convert_asset_manufacturer_info_db(self, asset):
        # 判空
        if asset.asset_manufacturer is None:
            return None
        # 数据转化为db对象
        asset_manufacturer_info_db = AssetManufacturesInfo(
            id=uuid.uuid4().hex,
            asset_id=asset.asset_id,
            name=asset.asset_manufacturer.name,
            description=asset.asset_manufacturer.description,
        )
        # 返回数据
        return asset_manufacturer_info_db


    # 创建资产与厂商关联关系
    def convert_manufacturer_relation_info_db(self, asset, manufacturer_info_db, manufacturer_info_db_new):
        # 判空
        if asset.asset_manufacturer is None:
            return None
        # 数据转化为db对象
        manufacturer_relation_info_db = AssetManufactureRelationInfo(
            id=uuid.uuid4().hex,
            asset_id=asset.asset_id,
            manufacture_id=None,
        )
        # 数据库中的厂商
        if manufacturer_info_db:
            manufacturer_relation_info_db.manufacture_id = manufacturer_info_db.id
        # 新建的厂商
        if manufacturer_info_db_new:
            manufacturer_relation_info_db.manufacture_id = manufacturer_info_db_new.id
        # 返回数据
        return manufacturer_relation_info_db

    # 资产创建时配件对象数据转换
    def convert_asset_part_info_db(self, asset):
        # 判空
        if asset.asset_part is None:
            return None
        # 配件列表
        asset_part_info_dbs = []
        # 遍历
        for temp in asset.asset_part:
            # 数据转化为db对象
            asset_part_info_db = AssetPartsInfo(
                id=uuid.uuid4().hex,
                asset_id=asset.asset_id,
                name=temp.name,
                part_type=temp.part_type,
                part_brand=temp.part_brand,
                part_config=temp.part_config,
                part_number=temp.part_number,
                personal_used_flag=temp.personal_used_flag,
                surplus=temp.surplus,
                description=temp.description,
            )
            #
            asset_part_info_dbs.append(asset_part_info_db)
        # 返回数据
        return asset_part_info_dbs


    # 资产创建时流量对象数据转换
    def convert_asset_flow_info_db_from_asset(self, asset):
        # 判空
        if asset.asset_flow is None:
            return None
        # 配件列表
        asset_flow_info_dbs = []
        # 遍历
        for temp in asset.asset_flow:
            # 数据转化为db对象
            asset_flow_info_db = AssetFlowsInfo(
                id=uuid.uuid4().hex,
                asset_id=asset.asset_id,
                port=asset.port,
                label=asset.label,
                opposite_asset_id=asset.opposite_asset_id,
                opposite_port=asset.opposite_label,
                opposite_label=asset.opposite_label,
                extra=json.dumps(temp.extra) if temp.extra else temp.extra,
                description=temp.description,
            )
            #
            asset_flow_info_dbs.append(asset_flow_info_db)
        # 返回数据
        return asset_flow_info_dbs


    # 资产创建时位置对象数据转换
    def convert_asset_position_info_db(self, asset):
        # 判空
        if asset.asset_position is None:
            return None
        # 数据转化为db对象
        asset_position_info_db = AssetPositionsInfo(
            id=uuid.uuid4().hex,
            asset_id=asset.asset_id,
            frame_position=asset.asset_position.frame_position,
            cabinet_position=asset.asset_position.cabinet_position,
            u_position=asset.asset_position.u_position,
            description=asset.asset_position.description,
        )
        # 返回数据
        return asset_position_info_db


    # 资产创建时合同对象数据转换
    def convert_asset_contract_info_db(self, asset):
        # 判空
        if asset.asset_contract is None:
            return None
        # 数据转化为db对象
        asset_contract_info_db = AssetContractsInfo(
            id=uuid.uuid4().hex,
            asset_id=asset.asset_id,
            contract_number=asset.asset_contract.contract_number,
            purchase_date=None if asset.asset_contract.purchase_date is None else datetime.fromtimestamp(asset.asset_contract.purchase_date/1000),
            batch_number=asset.asset_contract.batch_number,
            description=asset.asset_contract.description,
        )
        # 返回数据
        return asset_contract_info_db


    # 资产创建时合同对象数据转换
    def convert_asset_belong_info_db(self, asset):
        # 判空
        if asset.asset_belong is None:
            return None
        # 数据转化为db对象
        asset_contract_info_db = AssetBelongsInfo(
            id=uuid.uuid4().hex,
            asset_id=asset.asset_id,
            department_id=asset.asset_belong.department_id,
            department_name=asset.asset_belong.department_name,
            user_id=asset.asset_belong.user_id,
            user_name=asset.asset_belong.user_name,
            tel_number=asset.asset_belong.tel_number,
            description=asset.asset_belong.description,
        )
        # 返回数据
        return asset_contract_info_db


    # 资产创建时合同对象数据转换
    def convert_asset_customer_info_db(self, asset):
        # 判空
        if asset.asset_customer is None:
            return None
        # 数据转化为db对象
        asset_customer_info_db = AssetCustomersInfo(
            id=uuid.uuid4().hex,
            asset_id=asset.asset_id,
            customer_id=asset.asset_customer.customer_id,
            customer_name=asset.asset_customer.customer_name,
            rental_duration=asset.asset_customer.rental_duration,
            start_date=None if asset.asset_customer.start_date is None else datetime.fromtimestamp(asset.asset_customer.start_date/1000),
            end_date=None if asset.asset_customer.end_date is None else datetime.fromtimestamp(asset.asset_customer.end_date/1000),
            vlan_id=asset.asset_customer.vlan_id,
            float_ip=asset.asset_customer.float_ip,
            band_width=asset.asset_customer.band_width,
            description=asset.asset_customer.description,
        )
        # 返回数据
        return asset_customer_info_db

    def get_asset_by_id(self, asset_id):
        if not asset_id:
            return None
        # 详情
        try:
            # 根据id查询
            res = self.list_assets(asset_id, None,None, None,None,None, None, None, None, None, None, None, None, None, 1, 10, None, None)
            # 空
            if not res or not res.get("data"):
                return None
            # 返回第一条数据
            return res.get("data")[0]
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise Exception


    def update_assets_status(self, asset_list):
        # 列表空
        if not asset_list:
            return None
        # 详情
        try:
            for asset_temp in asset_list:
                # 判断对象是空
                if asset_temp is None:
                    pass
                # 资产id空或者资产状态空
                if not asset_temp.asset_id or not asset_temp.asset_status:
                    pass
                # 资产状态如果是错误的时候 描述信息不能为空
                if asset_temp.asset_status == "3" and not asset_temp.asset_status_description:
                    pass
                # 查询数据
                asset_basic_info_db = AssetSQL.get_asset_basic_info_by_id(asset_temp.asset_id)
                # 不存在
                if not asset_basic_info_db:
                    pass
                # 设置更新的字段
                asset_basic_info_db.asset_status = asset_temp.asset_status
                # 更新数据
                AssetSQL.update_asset_basic_info(asset_basic_info_db)
        except Exception as e:
            import traceback
            traceback.print_exc()
        # 成功返回资产id
        return None


    def delete_asset(self, asset_id):
        # 业务校验 todo
        if asset_id is None or len(asset_id) <= 0:
            return None
        # 删除
        try:
            # 删除对象
            AssetSQL.delete_asset(asset_id)
        except Exception as e:
            import traceback
            traceback.print_exc()
        # 成功返回资产id
        return None



    def import_asset(self, row):
        # 存入一行
        try:
            # 初始化数据对象
            asset = self.init_empty_asset_api_model()
            asset.asset_type_id = "8fb707d8-b07e-11ef-90c8-44a842237864"
            asset.asset_category = "SERVER"
            # 从excel中加载基础信息数据
            for basic_key, basic_column in asset_basic_info_columns.items():
                # 判断excel的数据是非nan
                if pd.notna(row[basic_column]):
                    asset.__setattr__(basic_key, row[basic_column])
            # 从excel中加载厂商信息数据
            for manufacture_key, manufacture_column in asset_manufacture_info_columns.items():
                # 判断excel的数据是非nan
                if pd.notna(row[manufacture_column]):
                    asset.asset_manufacturer.__setattr__(manufacture_key, row[manufacture_column])
            # 从excel中加载位置信息数据
            for position_key, position_column in asset_position_info_columns.items():
                # 判断excel的数据是非nan
                if pd.notna(row[position_column]):
                    asset.asset_position.__setattr__(position_key, row[position_column])
            # 从excel中加载合同信息数据
            for contract_key, contract_column in asset_contract_info_columns.items():
                # 判断excel的数据是非nan
                if pd.notna(row.get(contract_column, None)):
                    row_value = row[contract_column]
                    # 购买日期单独处理
                    if "purchase_date" == contract_key:
                        row_value = int(row_value.timestamp() * 1000)
                    # 赋值
                    asset.asset_contract.__setattr__(contract_key, row_value)
            # 从excel中加载归属信息数据
            for belong_key, belong_column in asset_belong_info_columns.items():
                # 判断excel的数据是非nan
                if pd.notna(row[belong_column]):
                    asset.asset_belong.__setattr__(belong_key, row[belong_column])
            # 从excel中加载租户信息数据
            for customer_key, customer_column in asset_customer_info_columns.items():
                # 判断excel的数据是非nan
                if pd.notna(row.get(customer_column, None)):
                    asset.asset_customer.__setattr__(customer_key, row.get(customer_column, None))
                    # 调用创建
            self.create_asset(asset)
            # 组装数据
        except Exception as e:
            import traceback
            traceback.print_exc()


    def import_asset_part(self, row):
        # 存入一行
        try:
            # 获取配件属于的资产编号
            asset_number = None
            if pd.notna(row["资产编号"]):
                asset_number = row["资产编号"]
            # 根据资产编号精准查询资产设备
            asset_id = None
            asset_basic_info_db = AssetSQL.get_asset_basic_info_by_asset_number(asset_number)
            if asset_basic_info_db:
                asset_id = asset_basic_info_db.id
            # 从excel中加载配件信息数据
            for part_key, part_column in asset_part_info_columns.items():
                # 每一列的数据都作为一个配件
                # 初始化数据对象
                asset_part = self.init_empty_asset_part_api_model()
                # 判断excel的数据是非nan
                if pd.notna(row[part_column]):
                    asset_part.__setattr__("part_type", part_key)
                    asset_part.__setattr__("asset_id", asset_id)
                    asset_part.__setattr__("name", row[part_column])
                # 入库
                AssetSQL.create_asset_part(asset_part)
        except Exception as e:
            import traceback
            traceback.print_exc()


    def import_asset_network(self, row):
        # 存入一行
        try:
            # 初始化数据对象
            asset = self.init_empty_asset_api_model()
            asset.asset_type_id = "8fbc77f1-b07e-11ef-90c8-44a842237864"
            asset.asset_category = "NETWORK"
            # 从网络设备的excel中加载基础信息数据
            for basic_key, basic_column in asset_network_basic_info_columns.items():
                # 判断excel的数据是非nan
                if pd.notna(row[basic_column]):
                    asset.__setattr__(basic_key, row[basic_column])
            # 从网络设备的excel中加载基础信息的扩展信息数据
            extra = {}
            for basic_extra_key, basic_extra_column in asset_network_basic_info_extra_columns.items():
                # 判断excel的数据是非nan
                if pd.notna(row[basic_extra_column]):
                    extra[basic_extra_key] = row[basic_extra_column]
            if extra:
                asset.extra = extra
            # 从网络设备的excel中加载厂商信息数据
            for manufacture_key, manufacture_column in asset_network_manufacture_info_columns.items():
                # 判断excel的数据是非nan
                if pd.notna(row[manufacture_column]):
                    asset.asset_manufacturer.__setattr__(manufacture_key, row[manufacture_column])
            # 从网络设备的excel中加载位置信息数据
            for position_key, position_column in asset_network_position_info_columns.items():
                # 判断excel的数据是非nan
                if pd.notna(row[position_column]):
                    asset.asset_position.__setattr__(position_key, row[position_column])
            # 从网络设备的excel中加载合同信息数据
            for contract_key, contract_column in asset_contract_info_columns.items():
                # 判断excel的数据是非nan
                if pd.notna(row.get(contract_column, None)):
                    row_value = row[contract_column]
                    # 购买日期单独处理
                    if "purchase_date" == contract_key:
                        row_value = int(row_value.timestamp() * 1000)
                    # 赋值
                    asset.asset_contract.__setattr__(contract_key, row_value)
            self.create_asset(asset)
            # 组装数据
        except Exception as e:
            import traceback
            traceback.print_exc()


    # 初始化一个空的资产数据的结构对象
    def init_empty_asset_api_model(self):
        # 初始化api的model对象
        asset = AssetCreateApiModel(
            # 资产通用信息
            asset_id = uuid.uuid4().hex,
            asset_name = "Default",
            asset_type_id = "Default",
            asset_description = None,
            # 资产基础信息
            equipment_number = None,
            sn_number = None,
            asset_number = None,
            asset_status = "0",
            extra = None,
            # 配件信息
            asset_part = None,
            # 位置信息
            asset_position = None,
            # 厂商信息
            asset_manufacturer = None,
            # 合同信息
            asset_contract = None,
            # 所属用户信息
            asset_belong = None,
            # 租户信息
            asset_customer = None,
        )
        # 厂商
        asset_manufacturer = AssetManufacturesInfo(
            id = uuid.uuid4().hex,
            asset_id = asset.asset_id,
            name = None,
            description = None,
            extra = None,
        )
        asset.asset_manufacturer = asset_manufacturer
        # 位置
        asset_position = AssetPositionsInfo(
            id = uuid.uuid4().hex,
            asset_id = asset.asset_id,
            frame_position = None,
            cabinet_position = None,
            u_position = None,
            description = None,
        )
        asset.asset_position = asset_position
        # 合同
        asset_contract = AssetContractsInfo(
            id = uuid.uuid4().hex,
            asset_id = asset.asset_id,
            contract_number = None,
            purchase_date = None,
            batch_number = None,
            description = None,
        )
        asset.asset_contract = asset_contract
        # 归属信息
        asset_belong = AssetBelongsInfo(
            id = uuid.uuid4().hex,
            asset_id = asset.asset_id,
            department_id = None,
            department_name = None,
            user_id = None,
            user_name = None,
            tel_number = None,
            description = None,
        )
        asset.asset_belong = asset_belong
        # 租户信息
        asset_customer = AssetCustomersInfo(
            id = uuid.uuid4().hex,
            asset_id = asset.asset_id,
            customer_id = None,
            customer_name = None,
            rental_duration = None,
            start_date = None,
            end_date = None,
            vlan_id = None,
            float_ip = None,
            band_width = None,
            description = None,
        )
        asset.asset_customer = asset_customer
        # 返回
        return asset

    # 初始化一个空的资产数据的结构对象
    def init_empty_asset_part_api_model(self):
        # 租户信息
        asset_part = AssetPartsInfo(
            id = uuid.uuid4().hex,
            asset_id = None,
            part_type = None,
            part_brand = None,
            part_config = None,
            part_number = None,
            personal_used_flag = None,
            surplus = None,
            name = None,
            description = None,
            extra = None,
        )
        # 返回
        return asset_part


    def create_asset_excel(self, asset_type, result_file_path):
        # 判空
        if not asset_type:
            return None
        # 模板路径
        current_template_file = None
        if asset_type == "server":
            current_template_file = os.getcwd() + ASSET_SERVER_TEMPLATE_FILE_DIR
        elif asset_type == "network":
            current_template_file = os.getcwd() + ASSET_NETWORK_TEMPLATE_FILE_DIR
        elif asset_type == "network_flow":
            current_template_file = os.getcwd() + ASSET_NETWORK_FLOW_TEMPLATE_FILE_DIR
        else:
            pass
        # 对应类型的模板不存在
        if current_template_file is None:
            return None
        # 生成不同资产类型的文件
        try:
            # 复制模板文件到临时目录
            shutil.copy2(current_template_file, result_file_path)
            # 服务器类型的文件
            if asset_type == "server":
                self.create_asset_server_excel(result_file_path)
            # 网络类型的文件
            elif asset_type == "network":
                self.create_asset_network_excel(result_file_path)
            # 网络类型流入流出的文件
            elif asset_type == "network_flow":
                self.create_asset_network_flow_excel(result_file_path, None)
            else:
                pass
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e


    def create_asset_server_excel(self, result_file_path):
        try:
            # 导出的excel数据
            excel_asset_data = []
            excel_part_data = []
            # 读取数据数据
            page = 1
            page_size = 100
            # 查询一页
            res = self.list_assets(None,None, None, "SERVER", None,None, None, None, None, None, None, None, None, None, page, page_size, None, None)
            while res and res['data']:
                # 写入数据
                for temp in res['data']:
                    # df.loc[index,"设备名称"] = temp["asset_name"]
                    # 下一行
                    # index = index + 1
                    # 修改或添加新数据
                    temp_asset_data = {'机架': temp['asset_position']['frame_position'],'机柜': temp['asset_position']['cabinet_position'],'U位': temp['asset_position']['u_position'],
                                       '设备名称': temp['asset_name'],'设备型号': temp['equipment_number'],'资产编号': temp['asset_number'],'序列号': temp['sn_number'],
                                       '部门': temp['asset_belong']['department_name'],'负责人': temp['asset_belong']['user_name'],'主机名': None,'IP': None,
                                       'IDRAC': None,'用途': None,'密码': None,'操作系统': None,
                                       '购买日期': temp['asset_contract']['purchase_date'],'厂商': temp['asset_manufacturer']['name'],'批次': temp['asset_contract']['batch_number'],'备注': temp['asset_description'],}
                    # 配件数据
                    if temp['asset_part']:
                        temp_part_data = {'资产编号': temp['asset_number'],}
                        # 遍历配件
                        for temp_part in temp['asset_part']:
                            # 表头存在
                            if asset_part_info_columns.get(temp_part['part_type'], None):
                                temp_part_data[asset_part_info_columns.get(temp_part['part_type'], None)] = temp_part['part_config']
                        # 数据追加
                        excel_part_data.append(temp_part_data)
                    # 加入excel导出数据列表
                    excel_asset_data.append(temp_asset_data)
                # 判断总页数已经大于等于当前已经查询的页数，已经查询完成，退出循环
                if res['totalPages'] >= page:
                    break
                # 查询下一页
                page = page + 1
                res = self.list_assets(None,None, None,None,None, None, None, None, None, None, None, None, None, None, page, page_size, None, None)
            # 加载模板文件
            book = load_workbook(result_file_path)
            sheet = book['asset']  # 默认使用第一个工作表
            # 确定起始写入行号
            start_row = 2  # 自动追加到最后一行之后
            # 写入数据到模板文件
            for idx, row in pd.DataFrame(excel_asset_data).iterrows():
                for col_idx, value in enumerate(row, start=1):  # 列从 1 开始
                    sheet.cell(row=start_row + idx, column=col_idx, value=value)
            # 保存修改
            book.save(result_file_path)
            # 配件sheet
            part_sheet = book['part']
            start_row = 2  # 自动追加到最后一行之后
            # 写入数据到模板文件
            for idx, row in pd.DataFrame(excel_part_data).iterrows():
                for col_idx, value in enumerate(row, start=1):  # 列从 1 开始
                    part_sheet.cell(row=start_row + idx, column=col_idx, value=value)
                    # 保存修改
            book.save(result_file_path)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e


    def create_asset_network_excel(self, result_file_path):
        try:
            # 导出的excel数据
            excel_asset_network_data = []
            # 读取数据数据
            page = 1
            page_size = 100
            # 查询一页
            res = self.list_assets(None,None, None, "NETWORK", None,None, None, None, None, None, None, None, None, None, page, page_size, None, None)
            while res and res['data']:
                # 写入数据
                for temp in res['data']:
                    # 修改或添加新数据
                    temp_asset_data = {'机柜': temp['asset_position']['cabinet_position'],'U位': temp['asset_position']['u_position'],
                                       '设备名称': temp['asset_name'],'设备型号': temp['equipment_number'],'资产编号': temp['asset_number'],
                                       '主机名': None,'管理地址': None,'带外网关': None,'m-lag mac': None,'网络设备角色': None,'序号': None,
                                       'loopback': None,'vlanifv4': None,'预留': None,'BGP_AS': None,'用途': None,
                                       '采购合同号': temp['asset_contract']['contract_number'],'厂商': temp['asset_manufacturer']['name'],}
                    # 加入excel导出数据列表
                    excel_asset_network_data.append(temp_asset_data)
                # 判断总页数已经大于等于当前已经查询的页数，已经查询完成，退出循环
                if res['totalPages'] >= page:
                    break
                # 查询下一页
                page = page + 1
                res = self.list_assets(None,None, None,None,None, None, None, None, None, None, None, None, None, None, page, page_size, None, None)
            # 加载模板文件
            book = load_workbook(result_file_path)
            sheet = book.active  # 默认使用第一个工作表
            # 确定起始写入行号
            start_row = 2  # 自动追加到最后一行之后
            # 写入数据到模板文件
            for idx, row in pd.DataFrame(excel_asset_network_data).iterrows():
                for col_idx, value in enumerate(row, start=1):  # 列从 1 开始
                    sheet.cell(row=start_row + idx, column=col_idx, value=value)
            # 保存修改
            book.save(result_file_path)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e


    def create_asset_network_excel_4select(self, result_file_path, ids):
        try:
            # 导出的excel数据
            excel_asset_network_data = []
            # 读取数据数据
            page = 1
            page_size = 100
            # 查询一页
            res = self.list_assets(None, ids, None, "NETWORK", None,None, None, None, None, None, None, None, None, None, page, page_size, None, None)
            while res and res['data']:
                # 写入数据
                for temp in res['data']:
                    # 修改或添加新数据
                    temp_asset_data = {'机柜': temp['asset_position']['cabinet_position'],'U位': temp['asset_position']['u_position'],
                                       '设备名称': temp['asset_name'],'设备型号': temp['equipment_number'],'资产编号': temp['asset_number'],
                                       '主机名': None,'管理地址': None,'带外网关': None,'m-lag mac': None,'网络设备角色': None,'序号': None,
                                       'loopback': None,'vlanifv4': None,'预留': None,'BGP_AS': None,'用途': None,
                                       '采购合同号': temp['asset_contract']['contract_number'],'厂商': temp['asset_manufacturer']['name'],}
                    # 加入excel导出数据列表
                    excel_asset_network_data.append(temp_asset_data)
                # 判断总页数已经大于等于当前已经查询的页数，已经查询完成，退出循环
                if res['totalPages'] >= page:
                    break
                # 查询下一页
                page = page + 1
                res = self.list_assets(None, ids, None,"NETWORK",None, None, None, None, None, None, None, None, None, None, page, page_size, None, None)
            # 加载模板文件
            book = load_workbook(result_file_path)
            sheet = book.active  # 默认使用第一个工作表
            # 确定起始写入行号
            start_row = 2  # 自动追加到最后一行之后
            # 写入数据到模板文件
            for idx, row in pd.DataFrame(excel_asset_network_data).iterrows():
                for col_idx, value in enumerate(row, start=1):  # 列从 1 开始
                    sheet.cell(row=start_row + idx, column=col_idx, value=value)
            # 保存修改
            book.save(result_file_path)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e


    def create_asset_excel_4batch(self, item, result_file_path):
        # 判空
        if not item or not item.asset_type or not item.asset_ids:
            return None
        # 模板路径
        current_template_file = None
        if item.asset_type == "server":
            current_template_file = os.getcwd() + ASSET_SERVER_TEMPLATE_FILE_DIR
        elif item.asset_type == "network":
            current_template_file = os.getcwd() + ASSET_NETWORK_TEMPLATE_FILE_DIR
        elif item.asset_type == "network_flow":
            current_template_file = os.getcwd() + ASSET_NETWORK_FLOW_TEMPLATE_FILE_DIR
        else:
            pass
        # 对应类型的模板不存在
        if current_template_file is None:
            return None
        # 生成不同资产类型的文件
        try:
            # 复制模板文件到临时目录
            shutil.copy2(current_template_file, result_file_path)
            # 服务器类型的文件
            if item.asset_type == "server":
                self.create_asset_excel_4select(result_file_path, item.asset_ids)
            # 网络类型的文件
            elif item.asset_type == "network":
                self.create_asset_network_excel_4select(result_file_path, item.asset_ids)
            # 网络类型流入流出的文件
            elif item.asset_type == "network_flow":
                self.create_asset_network_flow_excel(result_file_path, item.asset_ids)
            else:
                pass
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e


    def create_asset_excel_4select(self, result_file_path, ids):
        # 模板路径
        # template_file = os.getcwd() + ASSET_SERVER_TEMPLATE_FILE_DIR
        # 复制模板文件到临时目录
        # shutil.copy2(template_file, result_file_path)
        # 导出的excel数据
        excel_asset_data = []
        excel_part_data = []
        # 读取数据数据
        page = 1
        page_size = 100
        # 查询一页
        res = self.list_assets(None, ids,None, "SERVER", None,None, None, None, None, None, None, None, None, None, page, page_size, None, None)
        while res and res['data']:
            # 写入数据
            for temp in res['data']:
                # df.loc[index,"设备名称"] = temp["asset_name"]
                # 下一行
                # index = index + 1
                # 修改或添加新数据
                temp_asset_data = {'机架': temp['asset_position']['frame_position'],'机柜': temp['asset_position']['cabinet_position'],'U位': temp['asset_position']['u_position'],
                                   '设备名称': temp['asset_name'],'设备型号': temp['equipment_number'],'资产编号': temp['asset_number'],'序列号': temp['sn_number'],
                                   '部门': temp['asset_belong']['department_name'],'负责人': temp['asset_belong']['user_name'],'主机名': None,'IP': None,
                                   'IDRAC': None,'用途': None,'密码': None,'操作系统': None,
                                   '购买日期': temp['asset_contract']['purchase_date'],'厂商': temp['asset_manufacturer']['name'],'批次': temp['asset_contract']['batch_number'],'备注': temp['asset_description'],}
                # 配件数据
                if temp['asset_part']:
                    temp_part_data = {'资产编号': temp['asset_number'],}
                    # 遍历配件
                    for temp_part in temp['asset_part']:
                        # 表头存在
                        if asset_part_info_columns.get(temp_part['part_type'], None):
                            temp_part_data[asset_part_info_columns.get(temp_part['part_type'], None)] = temp_part['part_config']
                    # 数据追加
                    excel_part_data.append(temp_part_data)
                # 加入excel导出数据列表
                excel_asset_data.append(temp_asset_data)
            # 判断总页数已经大于等于当前已经查询的页数，已经查询完成，退出循环
            if res['totalPages'] >= page:
                break
            # 查询下一页
            page = page + 1
            res = self.list_assets(None, ids,None,"SERVER",None, None, None, None, None, None, None, None, None, None, page, page_size, None, None)
        try:
            # 加载模板文件
            book = load_workbook(result_file_path)
            sheet = book['asset']  # 默认使用第一个工作表
            # 确定起始写入行号
            start_row = 2  # 自动追加到最后一行之后
            # 写入数据到模板文件
            for idx, row in pd.DataFrame(excel_asset_data).iterrows():
                for col_idx, value in enumerate(row, start=1):  # 列从 1 开始
                    sheet.cell(row=start_row + idx, column=col_idx, value=value).border = thin_border
            # 保存修改
            # book.save(result_file_path)
            # 配件sheet
            part_sheet = book['part']
            start_row = 2  # 自动追加到最后一行之后
            # 写入数据到模板文件
            for idx, row in pd.DataFrame(excel_part_data).iterrows():
                for col_idx, value in enumerate(row, start=1):  # 列从 1 开始
                    part_sheet.cell(row=start_row + idx, column=col_idx, value=value).border = thin_border
                    # 保存修改
            book.save(result_file_path)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e

    # 批量更新
    def update_asset_list(self, asset_batch):
        try:
            # 判空
            if not asset_batch or not asset_batch.asset_ids:
                raise Exception
            # id分割
            asset_ids = asset_batch.asset_ids.split(',')
            # 判空
            if not asset_ids:
                raise Exception
            # 遍历
            for temp_asset_id in asset_ids:
                # 基础信息
                asset_basic_info_db = AssetSQL.get_asset_basic_info_by_id(temp_asset_id)
                # 判空
                if not asset_basic_info_db:
                    raise Exception
                # 临时更新对象
                temp_asset = AssetCreateApiModel(
                    asset_name=asset_basic_info_db.name,
                    asset_type_id=asset_basic_info_db.asset_type_id)
                # 设置更新参数
                if asset_batch.asset_type_id:
                    temp_asset.asset_type_id = asset_batch.asset_type_id
                if asset_batch.asset_type:
                    temp_asset.asset_type = asset_batch.asset_type
                # 更新
                self.update_asset(temp_asset_id, temp_asset)
            # 成功返回id
            return asset_ids
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e


    def update_asset(self, asset_id, asset):
        # 判空
        if not asset_id or not asset:
            raise Exception
        # 根据id数据
        asset_basic_info_db = AssetSQL.get_asset_basic_info_by_id(asset_id)
        # 基础数据不存在
        if asset_basic_info_db is None:
            raise Exception
        # 1、资产基础信息数据
        asset_basic_info_db = self.reset_asset_basic_info_db(asset_basic_info_db, asset)
        # 2、资产厂商信息 资产关联的厂商可能已经存在也可能不存在
        asset_manufacture_info_db = None
        asset_manufacture_info_db_new = None
        asset_manufacture_relation_info_db = None
        if asset.asset_manufacturer:
            asset_manufacture_info_db = self.check_manufacturer_exists(asset)
        # 数据库不存在当前厂商
        if not asset_manufacture_info_db:
            asset_manufacture_info_db_new = self.convert_asset_manufacturer_info_db(asset)
        # 建立资产与厂商关联关系对象
        asset_manufacture_relation_info_db = self.convert_manufacturer_relation_info_db(asset, asset_manufacture_info_db, asset_manufacture_info_db_new)
        # 3、资产位置信息
        asset_position_info_db = None
        if asset.asset_position:
            # 根据资产id查询
            asset_position_info_db = AssetSQL.get_position_by_asset_id(asset_id)
            if asset_position_info_db:
                asset_position_info_db = self.reset_asset_position_info_db(asset_position_info_db, asset)
            else:
                asset_position_info_db = self.convert_asset_position_info_db(asset)
        # 4、资产合同信息
        asset_contract_info_db = None
        if asset.asset_contract:
            # 根据资产id查询
            asset_contract_info_db = AssetSQL.get_contract_by_asset_id(asset_id)
            if asset_contract_info_db:
                asset_contract_info_db = self.reset_asset_contract_info_db(asset_contract_info_db, asset)
            else:
                asset_contract_info_db = self.convert_asset_contract_info_db(asset)
        # 5、资产归属信息
        asset_belong_info_db = None
        if asset.asset_belong:
            # 根据资产id查询
            asset_belong_info_db = AssetSQL.get_belong_by_asset_id(asset_id)
            if asset_belong_info_db:
                asset_belong_info_db = self.reset_asset_belong_info_db(asset_belong_info_db, asset)
            else:
                asset_belong_info_db = self.convert_asset_belong_info_db(asset)
        # 6、资产租赁信息
        asset_customer_info_db = None
        if asset.asset_customer:
            # 根据资产id查询
            asset_customer_info_db = AssetSQL.get_customer_by_asset_id(asset_id)
            if asset_customer_info_db:
                asset_customer_info_db = self.reset_asset_customer_info_db(asset_customer_info_db, asset)
            else:
                asset_customer_info_db = self.convert_asset_customer_info_db(asset)
        # 7、资产配件信息
        asset_part_info_db = None
        # 配件是列表重新保存
        if asset.asset_part:
            asset_part_info_db = self.convert_asset_part_info_db(asset)
        # 更新资产相关的数据
        AssetSQL.update_asset(asset_basic_info_db, asset_manufacture_info_db_new, asset_manufacture_relation_info_db, asset_position_info_db, asset_contract_info_db, asset_belong_info_db, asset_customer_info_db, asset_part_info_db, None)


    # 资产创建时基础对象数据转换
    def reset_asset_basic_info_db(self, asset_basic_info_db, asset):
        # 更新存在的字段
        if asset.asset_type_id:
            asset_basic_info_db.asset_type_id = asset.asset_type_id
        if asset.asset_category:
            asset_basic_info_db.asset_category = asset.asset_category
        if asset.asset_type:
            asset_basic_info_db.asset_type = asset.asset_type
        if asset.asset_name:
            asset_basic_info_db.name = asset.asset_name
        if asset.asset_description:
            asset_basic_info_db.description = asset.asset_description
        if asset.equipment_number:
            asset_basic_info_db.equipment_number = asset.equipment_number
        if asset.sn_number:
            asset_basic_info_db.sn_number = asset.sn_number
        if asset.asset_number:
            asset_basic_info_db.asset_number = asset.asset_number
        if asset.asset_status:
            asset_basic_info_db.asset_status = asset.asset_status
        if asset.extra:
            asset_basic_info_db.asset_status = json.dumps(asset.extra)
        # 返回数据
        return asset_basic_info_db


    # 重新设置位置信息
    def reset_asset_position_info_db(self, asset_position_info_db, asset):
        # 判空
        if asset.asset_position is None:
            return asset_position_info_db
        # 参数重设
        if asset.asset_position.frame_position:
            asset_position_info_db.frame_position=asset.asset_position.frame_position
        if asset.asset_position.cabinet_position:
            asset_position_info_db.cabinet_position=asset.asset_position.cabinet_position
        if asset.asset_position.u_position:
            asset_position_info_db.u_position=asset.asset_position.u_position
        if asset.asset_position.description:
            asset_position_info_db.description=asset.asset_position.description
        # 返回
        return asset_position_info_db

    # 重新设置合同信息
    def reset_asset_contract_info_db(self, asset_contract_info_db, asset):
        # 判空
        if asset.asset_contract is None:
            return asset_contract_info_db
        # 参数重设
        if asset.asset_contract.contract_number:
            asset_contract_info_db.contract_number=asset.asset_contract.contract_number
        if asset.asset_contract.purchase_date and asset.asset_contract.purchase_date > 0:
            asset_contract_info_db.purchase_date=datetime.fromtimestamp(asset.asset_contract.purchase_date/1000)
        if asset.asset_contract.batch_number:
            asset_contract_info_db.batch_number=asset.asset_contract.batch_number
        if asset.asset_contract.description:
            asset_contract_info_db.description=asset.asset_contract.description
        # 返回
        return asset_contract_info_db

    # 重新设置归属信息
    def reset_asset_belong_info_db(self, asset_belong_info_db, asset):
        # 判空
        if asset.asset_belong is None:
            return asset_belong_info_db
        # 参数重设
        if asset.asset_belong.department_id:
            asset_belong_info_db.department_id=asset.asset_belong.department_id
        if asset.asset_belong.department_name:
            asset_belong_info_db.department_name=asset.asset_belong.department_name
        if asset.asset_belong.user_id:
            asset_belong_info_db.user_id=asset.asset_belong.user_id
        if asset.asset_belong.user_name:
            asset_belong_info_db.user_name=asset.asset_belong.user_name
        if asset.asset_belong.tel_number:
            asset_belong_info_db.tel_number=asset.asset_belong.tel_number
        if asset.asset_belong.description:
            asset_belong_info_db.description=asset.asset_belong.description
        # 返回
        return asset_belong_info_db

        # 重新设置归属信息
    def reset_asset_customer_info_db(self, asset_customer_info_db, asset):
        # 判空
        if asset.asset_customer is None:
            return asset_customer_info_db
        # 参数重设
        if asset.asset_customer.customer_id:
            asset_customer_info_db.customer_id=asset.asset_customer.customer_id
        if asset.asset_customer.customer_name:
            asset_customer_info_db.customer_name=asset.asset_customer.customer_name
        if asset.asset_customer.rental_duration:
            asset_customer_info_db.rental_duration=asset.asset_customer.rental_duration
        if asset.asset_customer.start_date:
            asset_customer_info_db.start_date=datetime.fromtimestamp(asset.asset_customer.start_date/1000)
        if asset.asset_customer.end_date:
            asset_customer_info_db.end_date=datetime.fromtimestamp(asset.asset_customer.end_date/1000)
        if asset.asset_customer.vlan_id:
            asset_customer_info_db.vlan_id=asset.asset_customer.vlan_id
        if asset.asset_customer.float_ip:
            asset_customer_info_db.float_ip=asset.asset_customer.float_ip
        if asset.asset_customer.band_width:
            asset_customer_info_db.band_width=asset.asset_customer.band_width
        if asset.asset_customer.description:
            asset_customer_info_db.description=asset.asset_customer.description
        # 返回
        return asset_customer_info_db

    # 以下厂商相关功能 创建、查询、删除、修改
    def create_manufacture(self, manufacture):
        # 业务校验 todo
        if manufacture is None:
            return None
        # 定义id
        manufacture_id = None
        try:
            # 重名校验
            if manufacture.name is not None:
                count, _ = AssetSQL.list_manufacture(manufacture.name)
                if count > 0:
                    raise Exception
            # 数据组装
            # 资产厂商信息
            manufacture_info_db = self.convert_manufacturer_info_db(manufacture)
            manufacture_id = manufacture_info_db.id
            # 保存对象
            AssetSQL.create_manufacture(manufacture_info_db)
        except Exception as e:
            import traceback
            traceback.print_exc()
        # 成功返回资产id
        return manufacture_id

    # 创建时厂商对象数据转换
    def convert_manufacturer_info_db(self, manufacture):
        # 判空
        if manufacture is None:
            return None
        # 数据转化为db对象
        manufacturer_info_db = AssetManufacturesInfo(
            id=uuid.uuid4().hex,
            asset_id=manufacture.asset_id,
            name=manufacture.name,
            description=manufacture.name,
            extra=json.dumps(manufacture.extra) if manufacture.extra else manufacture.extra
        )
        # 返回数据
        return manufacturer_info_db

    # 查询厂商列表
    def list_manufactures(self, manufacture_name, page, page_size, sort_keys, sort_dirs):
        # 业务逻辑
        try:
            # 按照条件从数据库中查询数据
            count, data = AssetSQL.list_manufacture(manufacture_name, page, page_size, sort_keys, sort_dirs)
            # 数据处理
            ret = []
            # 遍历
            for r in data:
                # 填充数据 厂商信息
                temp_manufacture = {}
                temp_manufacture["id"] = r.id
                temp_manufacture["asset_id"] = r.asset_id
                temp_manufacture["name"] = r.name
                temp_manufacture["extra"] = r.extra
                temp_manufacture["description"] = r.description
                # 加入列表
                ret.append(temp_manufacture)
            # 返回数据
            res = {}
            # 页数相关信息
            if page and page_size:
                res['currentPage'] = page
                res['pageSize'] = page_size
                res['totalPages'] = ceil(count / int(page_size))
            res['total'] = count
            res['data'] = ret
            return res
        except Exception as e:
            import traceback
            traceback.print_exc()
            return None

    # 删除厂商
    def delete_manufacture(self, manufacture_id):
        # 业务校验 todo
        if manufacture_id is None or len(manufacture_id) <= 0:
            return None
        # 删除
        try:
            # 删除对象
            AssetSQL.delete_manufacture(manufacture_id)
        except Exception as e:
            import traceback
            traceback.print_exc()
        # 成功返回资产id
        return manufacture_id

    # 根据id修改厂商
    def update_manufacture(self, manufacture_id, manufacture_update_info):
        # 业务校验 todo
        if manufacture_update_info is None or manufacture_id is None or len(manufacture_id) <= 0:
            return None
        try:
            # 资产厂商信息
            manufacture_db = AssetSQL.get_manufacture_by_id(manufacture_id)
            # 判空
            if manufacture_db is None:
                return None
            # 填充需要修改的数据
            # 名称
            if manufacture_update_info.name is not None and len(manufacture_update_info.name) > 0:
                # 重名校验 名称不是当前名称 查询是否与其他名称重复
                if manufacture_update_info.name != manufacture_db.name:
                    count, _ = AssetSQL.list_manufacture(manufacture_update_info.name)
                    if count > 0:
                        raise Exception
                manufacture_db.name = manufacture_update_info.name
            # 描述
            if manufacture_update_info.description is not None and len(manufacture_update_info.description) > 0:
                manufacture_db.description = manufacture_update_info.description
            # 描述
            if manufacture_update_info.extra:
                manufacture_db.extra = json.dumps(manufacture_update_info.extra)
            # 保存对象
            AssetSQL.update_manufacture(manufacture_db)
        except Exception as e:
            import traceback
            traceback.print_exc()
        # 成功返回资产id
        return manufacture_id


# 以下是资产类型相关的service
    # 查询资产类型列表
    def list_assets_types(self, id, asset_type_name, asset_type_name_zh, child_included):
        # 业务逻辑
        try:
            # 按照条件从数据库中查询数据
            data = AssetSQL.list_asset_type(id, None,asset_type_name, asset_type_name_zh)
            # 数据处理
            ret = []
            # 遍历
            for r in data:
                # 填充数据
                temp = {}
                temp["id"] = r.id
                temp["parent_id"] = r.parent_id
                temp["asset_type_name"] = r.asset_type_name
                temp["asset_type_name_zh"] = r.asset_type_name_zh
                temp["queue"] = r.queue
                temp["description"] = r.description
                # 加入列表
                ret.append(temp)
            # 需要子级数据
            if child_included:
                pass
            # 返回数据
            return ret
        except Exception as e:
            import traceback
            traceback.print_exc()
            return None


    # 创建资产类型
    def create_asset_type(self, asset_type_api_model):
        # 业务逻辑
        try:
            # 业务校验 对象非空，数据线填充
            # 判空
            if asset_type_api_model is None or asset_type_api_model.asset_type_name is None:
                raise Exception
            # 数据对象转换
            asset_type_db = self.convert_asset_type_info_db(asset_type_api_model)
            # 父id校验
            parent_asset_type_db = None
            if asset_type_db.parent_id:
                # 查询父id的类型数据
                data = AssetSQL.list_asset_type(asset_type_db.parent_id, None,None, None)
                # 如果数据不存在
                if not data:
                    LOG.error("parent not exist")
                    raise Exception
            if asset_type_db.asset_type_name:
                # 查询名称重复的类型数据
                data = AssetSQL.list_asset_type(None, None, asset_type_db.asset_type_name, None)
                # 如果数据已经存在
                if data:
                    for temp_asset_type in data:
                        if asset_type_db.asset_type_name == temp_asset_type.asset_type_name:
                            LOG.error("asset type exist")
                            raise Exception
                # 获取数据
                # parent_asset_type_db = data[0]
                # 英文名称匹配校验
                # if not asset_type_db.asset_type_name.startswith(parent_asset_type_db.asset_type_name):
                #     LOG.error("asset type name is incompatible")
                #     raise Exception
            # 数据入库
            AssetSQL.create_asset_type(asset_type_db)
            # 返回
            return asset_type_db.id
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e

    # api的model对象转换数据库对象
    def convert_asset_type_info_db(self, asset_type_api_model):
        # 判空
        if asset_type_api_model is None:
            return None
        # 数据转化为db对象
        asset_type_db = AssetType(
            id=uuid.uuid4().hex,
            parent_id=asset_type_api_model.parent_id,
            asset_type_name=asset_type_api_model.asset_type_name,
            asset_type_name_zh=asset_type_api_model.asset_type_name_zh,
            queue=asset_type_api_model.queue,
            description=asset_type_api_model.description,
        )
        # 返回数据
        return asset_type_db

    # 查询资产类型及其子类型列表
    def list_child_asset_types(self, id):
        # 业务逻辑
        try:
            # 返回数据
            res = list()
            # 按照条件从数据库中查询数据
            data = AssetSQL.list_asset_type(None, id,None, None)
            # 数据处理
            if not data:
                return res
            # 加入当前数据
            res.extend(data)
            # 遍历
            for temp_asset_type in data:
                # 查询当前的子数据
                temp_data = self.list_child_asset_types(temp_asset_type.id)
                # 存在子数据
                if temp_data:
                    res.extend(temp_data)
            # 返回数据结果
            return res
        except Exception as e:
            import traceback
            traceback.print_exc()
            return None

    # 删除资产类型
    def delete_asset_type_by_id(self, asset_type_id):
        # 业务校验
        if asset_type_id is None or len(asset_type_id) <= 0:
            return None
        # 删除
        try:
            # 先删除下级然后再删除上级
            # 递归查询当前id的所有子类型
            child_data = self.list_child_asset_types(asset_type_id)
            # 遍历删除子类型
            if child_data:
                for temp_child_type in child_data:
                    # 删除子对象
                    AssetSQL.delete_asset_type(temp_child_type.id)
            # 删除对象
            AssetSQL.delete_asset_type(asset_type_id)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e
        # 成功返回资产id
        return asset_type_id


    # 根据id修改资产类型
    def update_asset_type_by_id(self, id, asset_type):
        # 业务校验
        if asset_type is None or id is None or len(id) <= 0:
            return None
        try:
            # 资产类型信息
            assert_type_db = AssetSQL.get_asset_type_by_id(id)
            # 判空
            if assert_type_db is None:
                return None
            # 填充需要修改的数据
            # 英文名称
            if asset_type.asset_type_name is not None and len(asset_type.asset_type_name) > 0:
                # 名称重复校验
                if asset_type.asset_type_name != assert_type_db.asset_type_name:
                    # 查询名称重复的类型数据
                    data = AssetSQL.list_asset_type(None, None, asset_type.asset_type_name, None)
                    # 如果数据已经存在
                    if data:
                        for temp_asset_type in data:
                            if asset_type.asset_type_name == temp_asset_type.asset_type_name:
                                LOG.error("asset type exist")
                                raise Exception
                # 校验通过赋值
                assert_type_db.asset_type_name = asset_type.asset_type_name
            # 描述
            if asset_type.description is not None and len(asset_type.description) > 0:
                assert_type_db.description = asset_type.description
            # 保存对象
            AssetSQL.update_asset_type(assert_type_db)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e
        # 成功返回id
        return id


    # 以下是资产配件相关方法
    # 查询资产配件列表
    def list_assets_parts(self, asset_id):
        # 业务逻辑
        try:
            # 按照条件从数据库中查询数据
            data = AssetSQL.list_asset_part(asset_id)
            # 数据处理
            ret = []
            # 遍历
            for r in data:
                # 填充数据
                temp = {}
                temp["id"] = r.id
                temp["name"] = r.name
                temp["asset_id"] = r.asset_id
                temp["part_type"] = r.part_type
                temp["part_brand"] = r.part_brand
                temp["part_config"] = r.part_config
                temp["part_number"] = r.part_number
                temp["personal_used_flag"] = r.personal_used_flag
                temp["surplus"] = r.surplus
                temp["description"] = r.description
                # 加入列表
                ret.append(temp)
            # 返回数据
            return ret
        except Exception as e:
            import traceback
            traceback.print_exc()
            return None

    # 查询资产配件列表
    def list_assets_parts_pages(self, part_catalog, asset_id, name, page, page_size, sort_keys, sort_dirs):
        # 业务逻辑
        try:
            # 按照条件从数据库中查询数据
            count, data = AssetSQL.list_asset_part_page(part_catalog, asset_id, name, page, page_size, sort_keys, sort_dirs)
            # 数据处理
            ret = []
            # 遍历
            for r in data:
                # 填充数据
                temp = {}
                temp["id"] = r.id
                temp["name"] = r.name
                temp["asset_name"] = r.asset_name
                temp["asset_number"] = r.asset_number
                temp["asset_id"] = r.asset_id
                temp["part_type"] = r.part_type
                temp["part_brand"] = r.part_brand
                temp["part_config"] = r.part_config
                temp["part_number"] = r.part_number
                temp["personal_used_flag"] = r.personal_used_flag
                temp["surplus"] = r.surplus
                temp["description"] = r.description
                # 加入列表
                ret.append(temp)
            # 返回数据
            res = {}
            # 页数相关信息
            if page and page_size:
                res['currentPage'] = page
                res['pageSize'] = page_size
                res['totalPages'] = ceil(count / int(page_size))
            res['total'] = count
            res['data'] = ret
            return res
        except Exception as e:
            import traceback
            traceback.print_exc()
            return None

    def create_asset_part(self, asset_part):
        # 判空
        if asset_part is None:
            return None
        # 定义id
        asset_part_id = None
        try:
            # 数据组装
            # 配件信息
            asset_part_info_db = self.convert_asset_part_info_db_4api(asset_part)
            asset_part_id = asset_part_info_db.id
            # 保存对象
            AssetSQL.create_asset_part(asset_part_info_db)
        except Exception as e:
            import traceback
            traceback.print_exc()
        # 成功返回资产id
        return asset_part_id


    # 创建配件对象数据转换 1个数据对象
    def convert_asset_part_info_db_4api(self, asset_part):
        # 判空
        if asset_part is None:
            return None
        # 数据转化为db对象
        asset_part_info_db = AssetPartsInfo(
            id=uuid.uuid4().hex,
            asset_id=asset_part.asset_id,
            name=asset_part.name,
            part_type=asset_part.part_type,
            part_brand=asset_part.part_brand,
            part_config=asset_part.part_config,
            part_number=asset_part.part_number,
            personal_used_flag=asset_part.personal_used_flag,
            surplus=asset_part.surplus,
            description=asset_part.description,
        )
        # 返回数据
        return asset_part_info_db

    # 根据id修改配件
    def update_asset_part_by_id(self, id, asset_part):
        # 业务校验
        if asset_part is None or id is None or len(id) <= 0:
            return None
        try:
            # 配件信息
            asset_part_db = AssetSQL.get_asset_part_by_id(id)
            # 判空
            if asset_part_db is None:
                return None
            # 填充需要修改的数据
            # 名称
            if asset_part.name is not None and len(asset_part.name) > 0:
                asset_part_db.name = asset_part.name
            # 类型
            if asset_part.part_type is not None and len(asset_part.part_type) > 0:
                asset_part_db.part_type = asset_part.part_type
            # 品牌
            if asset_part.part_brand is not None and len(asset_part.part_brand) > 0:
                asset_part_db.part_brand = asset_part.part_brand
            # 配置
            if asset_part.part_config is not None and len(asset_part.part_config) > 0:
                asset_part_db.part_config = asset_part.part_config
            # 编号
            if asset_part.part_number is not None and len(asset_part.part_number) > 0:
                asset_part_db.part_number = asset_part.part_number
            # 是否可自用
            if asset_part.personal_used_flag is not None:
                asset_part_db.personal_used_flag = asset_part.personal_used_flag
            # 剩余情况
            if asset_part.surplus is not None and len(asset_part.surplus) > 0:
                asset_part_db.surplus = asset_part.surplus
            # 描述
            if asset_part.description is not None and len(asset_part.description) > 0:
                asset_part_db.description = asset_part.description
            # 保存对象
            AssetSQL.update_asset_part(asset_part_db)
        except Exception as e:
            import traceback
            traceback.print_exc()
        # 成功返回资产id
        return id

    # 删除配件
    def delete_asset_part_by_id(self, asset_part_id):
        # 业务校验
        if asset_part_id is None or len(asset_part_id) <= 0:
            return None
        # 删除
        try:
            # 删除对象
            AssetSQL.delete_asset_part(asset_part_id)
        except Exception as e:
            import traceback
            traceback.print_exc()
        # 成功返回资产id
        return asset_part_id


    # 根据id修改配件
    def bind_asset_part_by_id(self, id, asset_id):
        # 业务校验
        if asset_id is None or len(asset_id) <= 0 or id is None or len(id) <= 0:
            return None
        try:
            # 配件信息
            asset_part_db = AssetSQL.get_asset_part_by_id(id)
            # 判空
            if asset_part_db is None:
                return None
            # 资产信息
            asset_basic_info_db = AssetSQL.get_asset_basic_info_by_id(asset_id)
            if asset_basic_info_db is None:
                return None
            # 填充需要修改的数据
            # id
            asset_part_db.asset_id = asset_id
            # 保存对象
            AssetSQL.update_asset_part(asset_part_db)
        except Exception as e:
            import traceback
            traceback.print_exc()
        # 成功返回资产id
        return id

        # 根据id修改配件
    def unbind_asset_part_by_id(self, id, asset_id):
        # 业务校验
        if asset_id is None or len(asset_id) <= 0 or id is None or len(id) <= 0:
            return None
        try:
            # 配件信息
            asset_part_db = AssetSQL.get_asset_part_by_id(id)
            # 判空
            if asset_part_db is None:
                return None
            # 资产信息
            asset_basic_info_db = AssetSQL.get_asset_basic_info_by_id(asset_id)
            if asset_basic_info_db is None:
                return None
            # 填充需要修改的数据
            # id
            asset_part_db.asset_id = None
            # 保存对象
            AssetSQL.update_asset_part(asset_part_db)
        except Exception as e:
            import traceback
            traceback.print_exc()
        # 成功返回资产id
        return id


    # 以下是资产-网络设备流相关的service
    # 查询资产流量列表
    def list_assets_flows(self, asset_id, ids):
        # 业务逻辑
        try:
            # 按照条件从数据库中查询数据
            data = AssetSQL.list_asset_flow(asset_id, ids)
            # 数据处理
            ret = []
            # 遍历
            for r in data:
                # 填充数据
                temp = {}
                temp["id"] = r.id
                temp["asset_id"] = r.asset_id
                temp["asset_name"] = r.asset_name
                temp["cabinet_position"] = r.cabinet_position
                temp["u_position"] = r.u_position
                temp["port"] = r.port
                temp["label"] = r.label
                temp["opposite_asset_id"] = r.opposite_asset_id
                temp["opposite_asset_name"] = r.opposite_asset_name
                temp["opposite_cabinet_position"] = r.opposite_cabinet_position
                temp["opposite_u_position"] = r.opposite_u_position
                temp["opposite_port"] = r.opposite_port
                temp["opposite_label"] = r.opposite_label
                temp["cable_type"] = r.cable_type
                temp["cable_interface_type"] = r.cable_interface_type
                temp["cable_length"] = r.cable_length
                temp["extra"] = r.extra
                temp["description"] = r.description
                # 加入列表
                ret.append(temp)
            # 返回数据
            return ret
        except Exception as e:
            import traceback
            traceback.print_exc()
            return None


    # 创建网络设备的流转数据
    def create_asset_flow(self, asset_flow_api_model):
        # 业务逻辑
        try:
            # 业务校验 对象非空，数据线填充
            # 判空
            if asset_flow_api_model is None or asset_flow_api_model.port is None:
                raise Exception
            # 数据对象转换
            asset_flow_db = self.convert_asset_flow_info_db(asset_flow_api_model)
            # 数据入库
            AssetSQL.create_asset_flow(asset_flow_db)
            # 返回
            return asset_flow_db.id
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e

    # api的model对象转换数据库对象
    def convert_asset_flow_info_db(self, asset_flow_api_model):
        # 判空
        if asset_flow_api_model is None:
            return None
        # 数据转化为db对象
        asset_flow_db = AssetFlowsInfo(
            id=uuid.uuid4().hex,
            asset_id=asset_flow_api_model.asset_id,
            port=asset_flow_api_model.port,
            label=asset_flow_api_model.label,
            opposite_asset_id=asset_flow_api_model.opposite_asset_id,
            opposite_port=asset_flow_api_model.opposite_port,
            opposite_label=asset_flow_api_model.opposite_label,
            cable_type=asset_flow_api_model.cable_type,
            cable_interface_type=asset_flow_api_model.cable_interface_type,
            cable_length=asset_flow_api_model.cable_length,
            extra=json.dumps(asset_flow_api_model.extra) if asset_flow_api_model.extra else asset_flow_api_model.extra,
            description=asset_flow_api_model.description,
        )
        # 返回数据
        return asset_flow_db

    # 删除网络设备的流转数据
    def delete_asset_flow_by_id(self, asset_flow_id):
        # 业务校验
        if asset_flow_id is None or len(asset_flow_id) <= 0:
            raise Exception
        # 删除
        try:
            # 删除对象
            AssetSQL.delete_asset_flow(asset_flow_id)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e
        # 成功返回id
        return asset_flow_id

    # 根据id修改网络设备的流转数据
    def update_asset_flow_by_id(self, id, asset_flow):
        # 业务校验
        if asset_flow is None or id is None or len(id) <= 0:
            raise Exception
        try:
            # 网络设备的流转信息
            assert_flow_db = AssetSQL.get_asset_flow_by_id(id)
            # 判空
            if assert_flow_db is None:
                raise Exception
            # 填充需要修改的数据
            # 端口
            if asset_flow.port is not None and len(asset_flow.port) > 0:
                assert_flow_db.port = asset_flow.port
            # 标签
            if asset_flow.label is not None and len(asset_flow.label) > 0:
                assert_flow_db.label = asset_flow.label
            # 对端资产id
            if asset_flow.opposite_asset_id is not None and len(asset_flow.opposite_asset_id) > 0:
                assert_flow_db.opposite_asset_id = asset_flow.opposite_asset_id
            # 对端端口
            if asset_flow.opposite_port is not None and len(asset_flow.opposite_port) > 0:
                assert_flow_db.opposite_port = asset_flow.opposite_port
            # 对端标签
            if asset_flow.opposite_label is not None and len(asset_flow.opposite_label) > 0:
                assert_flow_db.opposite_label = asset_flow.label
            # 线缆类型
            if asset_flow.cable_type is not None and len(asset_flow.cable_type) > 0:
                assert_flow_db.cable_type = asset_flow.cable_type
            # 线缆接口类型
            if asset_flow.cable_interface_type is not None and len(asset_flow.cable_interface_type) > 0:
                assert_flow_db.cable_interface_type = asset_flow.cable_interface_type
            # 线缆长度
            if asset_flow.cable_length is not None:
                assert_flow_db.cable_length = asset_flow.cable_length
            # 描述
            if asset_flow.description is not None and len(asset_flow.description) > 0:
                assert_flow_db.description = asset_flow.description
            # 保存对象
            AssetSQL.update_asset_flow(assert_flow_db)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e
        # 成功返回id
        return id

    # 导出网络设备的流数据，或者指定id的流数据
    def create_asset_network_flow_excel(self, result_file_path, ids):
        try:
            # 导出的excel数据
            excel_asset_network_flow_data = []
            # 查询
            res = self.list_assets_flows(None, ids)
            # 空
            if not res:
                return None
            # 非空，遍历写入
            for temp in res:
                # 修改或添加新数据
                temp_asset_data = {'设备名称': temp['asset_name'],'U位': temp['u_position'],'机柜': temp['cabinet_position'],
                                   '端口': temp['port'],'对端设备名称': temp['opposite_asset_name'],'对端U位': temp['opposite_u_position'],
                                   '对端机柜': temp['opposite_cabinet_position'],'线缆类型': temp['cable_type'],'线缆接口类型': temp['cable_interface_type'],
                                   '线缆长度': temp['cable_interface_type'],'标签': temp['label'],'对端标签': temp['opposite_label'],'备注': temp['description']}
                # 加入excel导出数据列表
                excel_asset_network_flow_data.append(temp_asset_data)
            # 加载模板文件
            book = load_workbook(result_file_path)
            sheet = book.active  # 默认使用第一个工作表
            # 确定起始写入行号
            start_row = 2  # 自动追加到最后一行之后
            # 写入数据到模板文件
            for idx, row in pd.DataFrame(excel_asset_network_flow_data).iterrows():
                for col_idx, value in enumerate(row, start=1):  # 列从 1 开始
                    sheet.cell(row=start_row + idx, column=col_idx, value=value)
            # 保存修改
            book.save(result_file_path)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e


    def import_asset_network_flow(self, row):
        # 存入一行
        try:
            # 初始化数据对象
            asset_network_flow = self.init_empty_asset_network_flow_api_model()
            asset_name = None
            opposite_asset_name = None
            # 从网络设备的excel中加载基础信息数据
            for basic_key, basic_column in asset_network_flow_info_columns.items():
                # 判断excel的数据是非nan 并且对象中存在对应属性
                if pd.notna(row[basic_column]):
                    # 资产名称
                    if basic_column == "设备名称":
                        asset_name = row[basic_column]
                    # 对端资产名称
                    if basic_column == "对端设备名称":
                        opposite_asset_name = row[basic_column]
                    # 属性与excel中能够对应
                    if hasattr(asset_network_flow, basic_key):
                        asset_network_flow.__setattr__(basic_key, row[basic_column])
            # 读取对应名称的资产设备信息
            if asset_name:
                asset_db = AssetSQL.get_asset_basic_info_by_catalog_name("NETWORK", asset_name)
                if asset_db:
                    asset_network_flow.asset_id = asset_db.id
            if opposite_asset_name:
                asset_db = AssetSQL.get_asset_basic_info_by_catalog_name("NETWORK", opposite_asset_name)
                if asset_db:
                    asset_network_flow.opposite_asset_id = asset_db.id
            # 保存
            self.create_asset_flow(asset_network_flow)
            # 组装数据
        except Exception as e:
            import traceback
            traceback.print_exc()

    # 初始化一个空的资产数据的结构对象
    def init_empty_asset_network_flow_api_model(self):
        # 初始化api的model对象
        asset_network_flow_api_model = AssetFlowApiModel(
            # 资产通用信息
            id = uuid.uuid4().hex,
            asset_id = None,
            port = None,
            label = None,
            opposite_asset_id = None,
            opposite_port = None,
            opposite_label = None,
            cable_type = None,
            cable_interface_type = None,
            cable_length = None,
            extra = None,
            description = None,
        )
        return asset_network_flow_api_model